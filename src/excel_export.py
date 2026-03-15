"""
excel_export.py — Exports the screener results to a formatted Excel workbook.
Creates separate worksheets for comparable companies, M&A transaction comps,
and the DCF model output. Uses openpyxl for all styling.
Returns BytesIO (for Streamlit) or writes to file path (for CLI).
"""

import io
import os
import platform
import subprocess
from typing import Union

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E4057"
LIGHT_BLUE   = "DEEAF1"
LIGHT_GRAY   = "F5F5F5"
BEAR_COLOR   = "FCE4D6"
BASE_COLOR   = "DEEAF1"
BULL_COLOR   = "E2EFDA"
SUMMARY_COLOR = "FFF2CC"

# ---------------------------------------------------------------------------
# Openpyxl style factories (created fresh each call — openpyxl requires it)
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, italic=False, size=11, color="000000") -> Font:
    return Font(bold=bold, italic=italic, size=size, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def style_header_row(ws, row_num: int, num_cols: int,
                     bg_color: str, font_color: str = "FFFFFF") -> None:
    """Fill row_num cells 1..num_cols with bg_color, white-bold font."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill    = _fill(bg_color)
        cell.font    = _font(bold=True, color=font_color)
        cell.alignment = _center()
        cell.border  = _thin_border()


def _write_title(ws, title: str, subtitle: str, num_cols: int) -> None:
    """Write merged title row + subtitle row."""
    last_col = get_column_letter(num_cols)

    # Row 1 — title
    ws.merge_cells(f"A1:{last_col}1")
    cell = ws["A1"]
    cell.value     = title
    cell.fill      = _fill(DARK_BLUE)
    cell.font      = _font(bold=True, size=14, color="FFFFFF")
    cell.alignment = _center()
    ws.row_dimensions[1].height = 24

    # Row 2 — subtitle / description
    ws.merge_cells(f"A2:{last_col}2")
    cell = ws["A2"]
    cell.value     = subtitle
    cell.font      = _font(italic=True, size=10, color="595959")
    cell.alignment = _left()
    ws.row_dimensions[2].height = 18

    # Row 3 — blank spacer
    ws.row_dimensions[3].height = 6


def _autofit_columns(ws, min_width: int = 14, max_width: int = 36) -> None:
    """Estimate column widths from cell content, skipping merged cells."""
    for col_cells in ws.columns:
        max_len = 0
        # MergedCell objects have no .column int attr — use the first real cell
        first = next((c for c in col_cells if not isinstance(c, MergedCell)), None)
        if first is None:
            continue
        col_letter = get_column_letter(first.column)
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_width, max(min_width, max_len * 1.2))
        ws.column_dimensions[col_letter].width = width


def _is_summary_label(value) -> bool:
    return str(value) in ("Median", "25th Pct", "75th Pct")


# ---------------------------------------------------------------------------
# Sheet 1 — Comparable Companies
# ---------------------------------------------------------------------------

def _write_comps_sheet(ws, comps_df: pd.DataFrame,
                       rationale: str, company_description: str) -> None:
    cols = list(comps_df.columns)
    n    = len(cols)

    _write_title(ws, "M&A Comparables Analysis", company_description, n)

    # Row 4 — column headers
    style_header_row(ws, 4, n, MID_BLUE)
    for c_idx, col_name in enumerate(cols, 1):
        ws.cell(row=4, column=c_idx).value = col_name

    ws.freeze_panes = "A5"

    # Determine which columns need number formats
    dollar_cols  = {c_idx for c_idx, col in enumerate(cols, 1) if "($M)" in col}
    multiple_cols = {c_idx for c_idx, col in enumerate(cols, 1)
                    if col in ("EV/EBITDA", "EV/Revenue", "P/E")}

    # Data rows
    for r_idx, (_, row) in enumerate(comps_df.iterrows(), start=5):
        is_summary = _is_summary_label(row.iloc[0])
        is_even    = (r_idx % 2 == 0)
        row_bg     = SUMMARY_COLOR if is_summary else (LIGHT_GRAY if is_even else "FFFFFF")

        for c_idx, val in enumerate(row, 1):
            # Coerce numpy scalar types → Python natives so openpyxl formats them
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)

            cell        = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = _fill(row_bg)
            cell.font   = _font(bold=is_summary, color="1F3864" if is_summary else "000000")
            cell.border = _thin_border()
            cell.alignment = _center() if c_idx > 1 else _left()

            # Number formatting — skip "NM", blank, None, strings
            if val not in ("NM", "", None) and isinstance(val, (int, float)):
                if c_idx in dollar_cols:
                    cell.number_format = "#,##0.0"
                elif c_idx in multiple_cols:
                    cell.number_format = '0.0"x"'

    # Blank row then rationale
    rationale_row = r_idx + 2
    ws.merge_cells(f"A{rationale_row}:{get_column_letter(n)}{rationale_row}")
    cell = ws.cell(row=rationale_row, column=1, value=rationale)
    cell.font      = _font(italic=True, size=10, color="666666")
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[rationale_row].height = 60

    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 2 — Transaction Comps
# ---------------------------------------------------------------------------

_TXN_COLS = ["Target", "Acquirer", "Year", "Deal Size ($M)",
             "EV/Revenue", "EV/EBITDA", "Notes"]
_TXN_KEYS = ["target", "acquirer", "year", "deal_size_usd_m",
             "ev_revenue", "ev_ebitda", "notes"]


def _write_transactions_sheet(ws, transactions: list[dict],
                               company_description: str) -> None:
    n = len(_TXN_COLS)
    _write_title(ws, "M&A Transaction Comparables", company_description, n)

    style_header_row(ws, 4, n, MID_BLUE)
    for c_idx, col_name in enumerate(_TXN_COLS, 1):
        ws.cell(row=4, column=c_idx).value = col_name

    ws.freeze_panes = "A5"

    deal_col     = _TXN_COLS.index("Deal Size ($M)") + 1
    ev_rev_col   = _TXN_COLS.index("EV/Revenue")     + 1
    ev_ebitda_col= _TXN_COLS.index("EV/EBITDA")      + 1

    for r_idx, txn in enumerate(transactions, start=5):
        is_even = (r_idx % 2 == 0)
        row_bg  = LIGHT_GRAY if is_even else "FFFFFF"

        for c_idx, key in enumerate(_TXN_KEYS, 1):
            raw = txn.get(key, "")
            val = raw

            # Coerce numeric columns
            if c_idx == deal_col and raw not in ("NA", "", None):
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    val = raw
            elif c_idx in (ev_rev_col, ev_ebitda_col) and raw not in ("NA", "", None):
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    val = raw

            cell        = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = _fill(row_bg)
            cell.font   = _font()
            cell.border = _thin_border()
            cell.alignment = _center() if c_idx != len(_TXN_COLS) else _left()

            if isinstance(val, float):
                if c_idx == deal_col:
                    cell.number_format = "#,##0"
                elif c_idx in (ev_rev_col, ev_ebitda_col):
                    cell.number_format = '0.0"x"'

    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 3 — DCF Analysis
# ---------------------------------------------------------------------------

_SCENARIO_ROWS = [
    ("Enterprise Value ($M)",   "enterprise_value",    "#,##0.0"),
    ("Equity Value ($M)",        "equity_value",       "#,##0.0"),
    ('Implied Share Price ($)',  "implied_share_price", "#,##0.00"),
    ("Terminal Value ($M)",      "terminal_value",     "#,##0.0"),
    ("TV as % of EV",            "terminal_value_pct", '0.0"%"'),
]


def _write_dcf_sheet(ws, dcf_scenarios: dict, dcf_table: pd.DataFrame) -> None:
    proj_cols = list(dcf_table.columns) if not dcf_table.empty else []
    n_cols    = max(len(proj_cols) + 1, 5)  # at least 5 wide

    _write_title(ws, "DCF Analysis", "5-Year Discounted Cash Flow Model", n_cols)

    current_row = 4

    # --- 5-Year Projection section ---
    # style_header_row first (sets fill/font on all cols), then overwrite col-1 value
    style_header_row(ws, current_row, n_cols, MID_BLUE)
    cell = ws.cell(row=current_row, column=1, value="5-Year Projection ($M)")
    cell.font      = _font(bold=True, size=11, color="FFFFFF")
    cell.fill      = _fill(MID_BLUE)
    cell.alignment = _left()

    if not dcf_table.empty:
        # Year headers
        current_row += 1
        ws.cell(row=current_row, column=1, value="Line Item")
        ws.cell(row=current_row, column=1).font  = _font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=1).fill  = _fill(MID_BLUE)
        ws.cell(row=current_row, column=1).alignment = _left()
        for c_idx, year_col in enumerate(proj_cols, 2):
            cell = ws.cell(row=current_row, column=c_idx, value=year_col)
            cell.fill = _fill(MID_BLUE)
            cell.font = _font(bold=True, color="FFFFFF")
            cell.alignment = _center()

        # Data rows
        for r_offset, (label, row) in enumerate(dcf_table.iterrows()):
            current_row += 1
            is_fcf  = (label == "Free Cash Flow")
            is_pv   = (label == "PV of FCF")
            is_df   = (label == "Discount Factor")
            is_even = (r_offset % 2 == 0)
            row_bg  = LIGHT_BLUE if (is_fcf or is_pv) else (LIGHT_GRAY if is_even else "FFFFFF")

            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font      = _font(bold=(is_fcf or is_pv))
            cell.fill      = _fill(row_bg)
            cell.alignment = _left()
            cell.border    = _thin_border()

            for c_idx, val in enumerate(row, 2):
                # Coerce numpy scalars → Python natives
                if isinstance(val, np.integer):
                    val = int(val)
                elif isinstance(val, np.floating):
                    val = float(val)
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.fill      = _fill(row_bg)
                cell.font      = _font(bold=(is_fcf or is_pv))
                cell.alignment = _center()
                cell.border    = _thin_border()
                if isinstance(val, (int, float)):
                    cell.number_format = "0.0000" if is_df else "#,##0.0"

    # --- Scenario Analysis section (2-row gap) ---
    current_row += 3
    # style_header_row first, then write value so it's not lost
    style_header_row(ws, current_row, n_cols, MID_BLUE)
    cell = ws.cell(row=current_row, column=1, value="Scenario Analysis")
    cell.font      = _font(bold=True, size=11, color="FFFFFF")
    cell.fill      = _fill(MID_BLUE)
    cell.alignment = _left()

    current_row += 1
    scenario_headers = [("", "FFFFFF", "FFFFFF"),
                        ("Bear Case", BEAR_COLOR, "000000"),
                        ("Base Case", BASE_COLOR,  "000000"),
                        ("Bull Case", BULL_COLOR,  "000000")]

    for c_offset, (label, bg, fg) in enumerate(scenario_headers):
        cell = ws.cell(row=current_row, column=1 + c_offset, value=label)
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=True, color=fg)
        cell.alignment = _center()
        cell.border    = _thin_border()

    for r_offset, (row_label, key, num_fmt) in enumerate(_SCENARIO_ROWS):
        current_row += 1
        is_even = (r_offset % 2 == 0)
        row_bg  = LIGHT_GRAY if is_even else "FFFFFF"

        cell = ws.cell(row=current_row, column=1, value=row_label)
        cell.font = _font(bold=True); cell.fill = _fill(row_bg)
        cell.alignment = _left(); cell.border = _thin_border()

        for c_offset, scenario in enumerate(("bear", "base", "bull"), 1):
            scenario_bg = [BEAR_COLOR, BASE_COLOR, BULL_COLOR][c_offset - 1]
            val  = dcf_scenarios.get(scenario, {}).get(key)
            cell = ws.cell(row=current_row, column=1 + c_offset, value=val)
            cell.fill      = _fill(scenario_bg)
            cell.alignment = _center()
            cell.border    = _thin_border()
            if val is not None:
                cell.number_format = num_fmt

    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_to_excel(
    comps_df: pd.DataFrame,
    transactions_list: list[dict],
    dcf_scenarios: dict,
    dcf_table: pd.DataFrame,
    rationale: str,
    company_description: str,
    output: Union[str, None] = None,
) -> Union[io.BytesIO, str]:
    """
    Build and return a formatted Excel workbook.

    Parameters
    ----------
    comps_df            : DataFrame from build_comps_table()
    transactions_list   : list[dict] from find_transactions()
    dcf_scenarios       : dict from run_scenarios()
    dcf_table           : DataFrame from build_dcf_table()
    rationale           : str from generate_rationale()
    company_description : str  free-text description shown as subtitle
    output              : None -> return BytesIO
                          str  -> write file, return path

    Returns BytesIO or file path string. Returns None on failure.
    """
    try:
        wb = Workbook()

        # Sheet 1 — Comparable Companies
        ws1 = wb.active
        ws1.title = "Comparable Companies"
        _write_comps_sheet(ws1, comps_df, rationale, company_description)

        # Sheet 2 — Transaction Comps
        ws2 = wb.create_sheet("Transaction Comps")
        _write_transactions_sheet(ws2, transactions_list, company_description)

        # Sheet 3 — DCF Analysis
        ws3 = wb.create_sheet("DCF Analysis")
        _write_dcf_sheet(ws3, dcf_scenarios, dcf_table)

        # Save
        if output is None:
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf
        else:
            wb.save(output)
            return output

    except Exception as e:
        print(f"[WARNING] excel_export: export failed — {e}")
        import traceback; traceback.print_exc()
        return None


# ---------------------------------------------------------------------------
# Smoke-test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import os
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from src.dcf import build_dcf_table, run_scenarios
    from src.multiples import COLUMNS

    # --- Minimal comps DataFrame ---
    comps_data = [
        ["Workday",   "WDAY",      "US",    35535.0, 33913.0,  9552.0, 1371.0, 24.7, 3.6, 51.3],
        ["Paycom",    "PAYC",      "US",     6902.2,  6622.5,  2051.7,  634.5, 10.4, 3.2, 15.2],
        ["Info Edge", "NAUKRI.NS", "INDIA", 616679.3,603385.5,31653.5,10375.5, 58.2,19.1, 45.8],
        ["Median",    "",          "",      "",      "",      "",      "",      24.7, 3.6, 45.8],
        ["25th Pct",  "",          "",      "",      "",      "",      "",      17.6, 3.4, 30.5],
        ["75th Pct",  "",          "",      "",      "",      "",      "",      41.5,11.4, 48.5],
    ]
    comps_df = pd.DataFrame(comps_data, columns=COLUMNS)

    # --- Minimal transactions ---
    transactions = [
        {"target": "Darwinbox", "acquirer": "TPG Capital",
         "year": "2022", "deal_size_usd_m": "950",
         "ev_revenue": "15.0", "ev_ebitda": "NA",
         "notes": "Series E funding at unicorn valuation for India HR SaaS"},
        {"target": "Keka HR",   "acquirer": "WestBridge Capital",
         "year": "2022", "deal_size_usd_m": "57",
         "ev_revenue": "8.5",  "ev_ebitda": "NA",
         "notes": "Growth equity round for India payroll SaaS platform"},
        {"target": "Workday",   "acquirer": "Acquired HiredScore",
         "year": "2024", "deal_size_usd_m": "524",
         "ev_revenue": "NA",   "ev_ebitda": "NA",
         "notes": "AI-powered talent management bolt-on for Workday HCM"},
    ]

    # --- DCF ---
    dcf_params = {
        "base_revenue": 20, "growth_rates": [0.40, 0.35, 0.28, 0.22, 0.18],
        "ebitda_margin": 0.15, "wacc": 0.14, "terminal_growth_rate": 0.04,
        "net_debt": -5, "shares_outstanding": 10,
    }
    dcf_table     = build_dcf_table(dcf_params)
    dcf_scenarios = run_scenarios(dcf_params)

    rationale = (
        "These companies are selected as comparables because they share a "
        "B2B SaaS model focused on human capital management for mid-market "
        "enterprises. Each generates recurring subscription revenue with "
        "similar gross margin profiles and sales motion to the target company."
    )
    description = (
        "B2B SaaS HR & Payroll platform, India mid-market focus, ~$20M ARR, Series B"
    )

    out_path = "test_output.xlsx"
    result = export_to_excel(
        comps_df=comps_df,
        transactions_list=transactions,
        dcf_scenarios=dcf_scenarios,
        dcf_table=dcf_table,
        rationale=rationale,
        company_description=description,
        output=out_path,
    )

    if result:
        print(f"Exported to {result}")
        # Auto-open
        if platform.system() == "Windows":
            os.startfile(os.path.abspath(result))
        elif platform.system() == "Darwin":
            subprocess.run(["open", result])
        else:
            subprocess.run(["xdg-open", result])
    else:
        print("Export failed — check warnings above.")
